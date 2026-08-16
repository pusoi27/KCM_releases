#*****************************
#student_manager.py   ver 04--------------
#*****************************

import sqlite3, csv, os, json
from modules.database import DB_PATH, issue_unique_qr_token
from modules import qr_generator

MAX_SUBJECTS = 3
MAX_SCHEDULE_DAYS = 6


def _photo_url(student_id, has_photo):
    return f"/students/photo/{student_id}" if has_photo else ''


def _coerce_blob(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, memoryview):
        return raw_value.tobytes()
    if isinstance(raw_value, bytes):
        return raw_value
    return None


def _loads_json_list(raw_value, default=None):
    """Safely decode a JSON list value."""
    if default is None:
        default = []
    if not raw_value:
        return list(default)
    try:
        data = json.loads(raw_value)
    except (TypeError, ValueError):
        return list(default)
    return data if isinstance(data, list) else list(default)


def _classification_label(el=0, pi=0, v=0, ind=0, **_):
    """Return the primary classification label for a student."""
    if int(bool(el)):
        return "Assisted"
    if int(bool(pi)):
        return "Monitored"
    if int(bool(ind)):
        return "Independent"
    if int(bool(v)):
        return "Virtual"
    return "Monitored"


def _normalize_schedule_entries(schedule_json, *schedule_pairs):
    """Return normalized schedule entries from JSON with legacy fallback."""
    entries = []
    seen = set()

    for entry in _loads_json_list(schedule_json):
        if not isinstance(entry, dict):
            continue
        day = str(entry.get('day') or '').strip()
        time = str(entry.get('time') or '').strip()
        if not day or day in seen:
            continue
        seen.add(day)
        entries.append({'day': day, 'time': time})

    if not entries:
        for day, time in schedule_pairs:
            day = str(day or '').strip()
            if not day or day in seen:
                continue
            seen.add(day)
            entries.append({'day': day, 'time': str(time or '').strip()})

    return entries[:MAX_SCHEDULE_DAYS]


def _build_student_database_row(row):
    """Convert a database row into the student database view model."""
    subjects = [str(s).strip() for s in _loads_json_list(row[10]) if str(s or '').strip()]
    if not subjects and row[2]:
        subjects = [str(row[2]).strip()]

    subject_slots = [""] * MAX_SUBJECTS
    for idx, subject_name in enumerate(subjects[:MAX_SUBJECTS]):
        subject_slots[idx] = subject_name

    schedule_entries = _normalize_schedule_entries(
        row[11],
        (row[12], row[13]),
        (row[14], row[15]),
        (row[22] if len(row) > 22 else '', row[23] if len(row) > 23 else ''),
        (row[24] if len(row) > 24 else '', row[25] if len(row) > 25 else ''),
        (row[26] if len(row) > 26 else '', row[27] if len(row) > 27 else ''),
        (row[28] if len(row) > 28 else '', row[29] if len(row) > 29 else ''),
    )
    schedule_slots = [None] * MAX_SCHEDULE_DAYS
    for idx, entry in enumerate(schedule_entries[:MAX_SCHEDULE_DAYS]):
        schedule_slots[idx] = entry

    photo_blob = _coerce_blob(row[17] if len(row) > 17 else None)
    return {
        'id': row[0],
        'name': row[1],
        'student_identifier': str(row[30] or '') if len(row) > 30 else '',
        'subject': row[2],
        'email': row[3],
        'phone': row[4],
        'guardian': str(row[19] or '') if len(row) > 19 else '',
        'secondary_email': str(row[31] or '') if len(row) > 31 else '',
        'secondary_phone': str(row[32] or '') if len(row) > 32 else '',
        'secondary_guardian': str(row[33] or '') if len(row) > 33 else '',
        'active': bool(row[5]),
        'book_loaned': bool(row[6]),
        'device_loaned': bool(row[7]),
        'el': bool(row[8]),
        'pi': bool(row[9]),
        'subjects': subjects,
        'subject_slots': subject_slots,
        'photo_url': _photo_url(row[0], bool(photo_blob)),
        'has_photo': bool(photo_blob),
        'classification': _classification_label(el=row[8], pi=row[9], v=row[16], ind=row[20] if len(row) > 20 else 0),
        'virtual': bool(row[16]),
        'schedule': schedule_entries,
        'schedule_slots': schedule_slots,
        'qr_filename': f"student_{row[0]}.png",
        'checkout_notify_enabled': bool(row[21]) if len(row) > 21 else True,
    }


def get_student_database_rows(active=1):
    """Get student rows tailored for the Student Database screen only."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        cols = [row[1] for row in c.execute("PRAGMA table_info(students)").fetchall()]
        has_checkout_notify = "checkout_notify_enabled" in cols

        has_book_loans_table = c.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='book_loans' LIMIT 1"
        ).fetchone() is not None
        has_material_loans_table = c.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='material_loans' LIMIT 1"
        ).fetchone() is not None

        checkout_notify_select = "checkout_notify_enabled" if has_checkout_notify else "1 AS checkout_notify_enabled"
        book_loaned_select = (
            "CASE WHEN EXISTS ("
            "SELECT 1 FROM book_loans bl "
            "WHERE bl.student_id = students.id AND bl.return_date IS NULL"
            ") THEN 1 ELSE COALESCE(students.book_loaned, 0) END"
            if has_book_loans_table
            else "COALESCE(students.book_loaned, 0)"
        )
        device_loaned_select = (
            "CASE WHEN EXISTS ("
            "SELECT 1 FROM material_loans ml "
            "WHERE ml.student_id = students.id AND ml.return_date IS NULL"
            ") THEN 1 ELSE COALESCE(students.device_loaned, 0) END"
            if has_material_loans_table
            else "COALESCE(students.device_loaned, 0)"
        )
        query = (
            "SELECT "
            f"id, name, subject, email, phone, active, {book_loaned_select} AS book_loaned, {device_loaned_select} AS device_loaned, "
            "el, pi, subjects_json, schedule_json, day1, day1_time, day2, day2_time, "
            "v, photo_blob, photo_mime, guardian, ind, "
            f"{checkout_notify_select} "
            ", day3, day3_time, day4, day4_time, day5, day5_time, day6, day6_time, COALESCE(student_identifier, '') AS student_identifier, COALESCE(secondary_email, '') AS secondary_email, COALESCE(secondary_phone, '') AS secondary_phone, COALESCE(secondary_guardian, '') AS secondary_guardian "
            "FROM students "
            "WHERE active = ? "
            "ORDER BY name"
        )

        c.execute(query, (active,))
        return [_build_student_database_row(row) for row in c.fetchall()]


def safe_int(value, default=0):
    """Safely convert value to int, returning default if empty or invalid."""
    try:
        val = str(value).strip()
        if not val:
            return default
        return int(val)
    except (ValueError, TypeError):
        return default


def _is_repeating_digit_student_id(candidate_id):
    """Return True when a numeric ID has all identical digits (e.g., 11, 111, 2222)."""
    token = str(candidate_id or '').strip()
    return len(token) >= 2 and token.isdigit() and len(set(token)) == 1


def _next_student_id_candidate(conn):
    """Best-effort next student id prediction for AUTOINCREMENT-backed tables."""
    seq_row = conn.execute(
        "SELECT seq FROM sqlite_sequence WHERE name='students'",
        (),
    ).fetchone()
    if seq_row and seq_row[0] is not None:
        return int(seq_row[0]) + 1

    max_row = conn.execute("SELECT COALESCE(MAX(id), 0) FROM students", ()).fetchone()
    return int((max_row[0] if max_row else 0) or 0) + 1


def _advance_student_id_sequence_past_repeating_digits(conn):
    """Skip repeating-digit student IDs by advancing sqlite_sequence before insert.

    This keeps the QR payload's `ID:<sid>` value on non-repeating IDs while
    preserving the DB-primary-key-based ID model.
    """
    try:
        next_id = _next_student_id_candidate(conn)
        skipped = 0
        while _is_repeating_digit_student_id(next_id):
            # Move ahead by +1 until at least one digit differs from the rest.
            conn.execute(
                "INSERT INTO sqlite_sequence(name, seq) VALUES('students', ?) "
                "ON CONFLICT(name) DO UPDATE SET seq=excluded.seq",
                (next_id,),
            )
            skipped += 1
            next_id += 1

        if skipped > 0:
            print(f"[student-id] Skipped {skipped} repeating-digit ID(s); next student ID will start at {next_id}.")
    except sqlite3.OperationalError:
        # sqlite_sequence may be unavailable when table is not AUTOINCREMENT.
        pass


def _normalize_csv_key(key):
    return str(key or '').strip().lower().replace(' ', '').replace('_', '').replace('-', '')


def _csv_get(row, *aliases, default=''):
    """Case/format-insensitive CSV field lookup."""
    if not isinstance(row, dict):
        return default
    wanted = {_normalize_csv_key(a) for a in aliases if a}
    for key, value in row.items():
        if _normalize_csv_key(key) in wanted:
            return value if value is not None else default
    return default


def _csv_bool(value):
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'y', 'on'}


def _csv_marked(value):
    """Return True for marker-style CSV values (e.g., x)."""
    token = str(value or '').strip().lower()
    return token in {'x', '1', 'true', 'yes', 'y', 'on', '✓', '✔'}


def normalize_student_identifier(value, max_len=64):
    """Normalize student identifier input to an alphanumeric token."""
    cleaned = ''.join(ch for ch in str(value or '').strip() if ch.isalnum())
    if max_len and len(cleaned) > max_len:
        cleaned = cleaned[:max_len]
    return cleaned


def _normalize_subject_name(raw_subject):
    token = str(raw_subject or '').strip()
    key = token.lower()

    # Legacy S# subject designations are mapped without hard-coded legacy labels.
    if key.startswith('s') and key[1:].isdigit():
        if key[1:] == '1':
            return 'Reading'
        if key[1:] == '2':
            return 'Math'

    mapping = {
        'm': 'Math',
        'math': 'Math',
        'mathematics': 'Math',
        'r': 'Reading',
        'reading': 'Reading',
        'read': 'Reading',
        'w': 'Writing',
        'writing': 'Writing',
        'write': 'Writing',
    }
    return mapping.get(key, token)


def _parse_subjects_from_csv(row):
    """Parse subjects from CSV values (new and legacy formats)."""
    subjects = []

    # Preferred format: M/R/W columns, marker value means selected.
    if _csv_marked(_csv_get(row, 'm', 'math', default='')):
        subjects.append('Math')
    if _csv_marked(_csv_get(row, 'r', 'reading', default='')):
        subjects.append('Reading')
    if _csv_marked(_csv_get(row, 'w', 'writing', default='')):
        subjects.append('Writing')

    raw_subjects = str(_csv_get(row, 'subjects', 'subjects_json', default='') or '').strip()
    if raw_subjects and not subjects:
        if '|' in raw_subjects:
            tokens = raw_subjects.split('|')
        elif ';' in raw_subjects:
            tokens = raw_subjects.split(';')
        else:
            tokens = raw_subjects.split(',')
        subjects.extend([t.strip() for t in tokens if t and t.strip()])

    for key in ('subject_1', 'subject_2', 'subject_3'):
        val = str(_csv_get(row, key, default='') or '').strip()
        if val and not subjects:
            subjects.append(val)

    legacy_subject = str(_csv_get(row, 'subject', default='') or '').strip()
    if legacy_subject and not subjects:
        subjects = [legacy_subject]

    deduped = []
    seen = set()
    for s in subjects:
        normalized_subject = _normalize_subject_name(s)
        k = normalized_subject.lower().strip()
        if not k or k in seen:
            continue
        seen.add(k)
        deduped.append(normalized_subject)

    if not deduped:
        deduped = ['Math']

    return deduped[:MAX_SUBJECTS]


def _parse_classification_from_csv(row):
    """Parse classification into checkbox flags used by the edit form."""
    el = int(_csv_bool(_csv_get(row, 'el', 'assisted', default='0')))
    pi = int(_csv_bool(_csv_get(row, 'pi', 'monitored', default='0')))
    v = int(_csv_bool(_csv_get(row, 'v', 'virtual', default='0')))
    ind = int(_csv_bool(_csv_get(row, 'ind', 'independent', default='0')))

    raw = str(_csv_get(row, 'classification', default='') or '').strip().lower()
    if raw:
        if raw in ('assisted', 'a'):
            return 1, 0, 0, 0
        if raw in ('monitored', 'm'):
            return 0, 1, 0, 0
        if raw in ('independent', 'i'):
            return 0, 0, 0, 1
        if raw in ('virtual', 'v'):
            return 0, 0, 1, 0

    # Enforce single primary classification (same model as edit form)
    if el:
        return 1, 0, 0, 0
    if pi:
        return 0, 1, 0, 0
    if ind:
        return 0, 0, 0, 1
    if v:
        return 0, 0, 1, 0
    return 0, 1, 0, 0


def _normalize_schedule_entries_from_json(raw_schedule_json):
    entries = []
    seen = set()
    if not raw_schedule_json:
        return entries
    try:
        data = json.loads(raw_schedule_json)
    except (TypeError, ValueError):
        return entries
    if not isinstance(data, list):
        return entries
    for entry in data:
        if not isinstance(entry, dict):
            continue
        day = str(entry.get('day') or '').strip()
        time = str(entry.get('time') or '').strip()
        if not day or day in seen:
            continue
        seen.add(day)
        entries.append({'day': day, 'time': time})
        if len(entries) >= MAX_SCHEDULE_DAYS:
            break
    return entries


def _parse_schedule_from_csv(row):
    """Parse schedule in one of these formats:
    - schedule_json: JSON list of {day,time}
    - schedule: Monday@15:00|Wednesday@16:00 (also accepts `Day Time`)
    - legacy day1/day1_time/day2/day2_time columns
    """
    entries = _normalize_schedule_entries_from_json(_csv_get(row, 'schedule_json', default=''))

    raw_schedule = str(_csv_get(row, 'schedule', default='') or '').strip()
    if raw_schedule:
        seen = {e['day'] for e in entries}
        for chunk in raw_schedule.split('|'):
            token = str(chunk or '').strip()
            if not token:
                continue
            day = ''
            time = ''
            if '@' in token:
                day, time = token.split('@', 1)
            elif ':' in token and ' ' in token:
                day, time = token.rsplit(' ', 1)
            else:
                day = token
            day = str(day or '').strip()
            time = str(time or '').strip()
            if not day or day in seen:
                continue
            seen.add(day)
            entries.append({'day': day, 'time': time})
            if len(entries) >= MAX_SCHEDULE_DAYS:
                break

    if not entries:
        seen = set()
        for day_key, time_key in (
            ('day1', 'day1_time'),
            ('day2', 'day2_time'),
            ('day3', 'day3_time'),
            ('day4', 'day4_time'),
            ('day5', 'day5_time'),
            ('day6', 'day6_time'),
        ):
            day = str(_csv_get(row, day_key, default='') or '').strip()
            time = str(_csv_get(row, time_key, default='') or '').strip()
            if not day or day in seen:
                continue
            seen.add(day)
            entries.append({'day': day, 'time': time})

    entries = entries[:MAX_SCHEDULE_DAYS]
    slot_values = []
    for idx in range(MAX_SCHEDULE_DAYS):
        if idx < len(entries):
            slot_values.extend([entries[idx]['day'], entries[idx]['time']])
        else:
            slot_values.extend(['', ''])
    return (*slot_values, json.dumps(entries))


def normalize_subject_entries(subjects, minutes):
    """Normalize subjects and their durations.

    Returns:
        tuple[list[str], list[int], int]: (subjects, minutes, total_minutes)
    """
    cleaned_subjects = []
    cleaned_minutes = []

    for idx, raw_subj in enumerate(subjects or []):
        subj = str(raw_subj or "").strip()
        if not subj:
            continue
        minute_raw = minutes[idx] if idx < len(minutes or []) else 30
        minute_val = max(5, safe_int(minute_raw, default=30))
        cleaned_subjects.append(subj)
        cleaned_minutes.append(minute_val)

    if not cleaned_subjects:
        cleaned_subjects = ["Math"]
        cleaned_minutes = [30]

    cleaned_subjects = cleaned_subjects[:MAX_SUBJECTS]
    cleaned_minutes = cleaned_minutes[:MAX_SUBJECTS]

    total_minutes = sum(cleaned_minutes) if cleaned_minutes else 30
    return cleaned_subjects, cleaned_minutes, total_minutes


def get_all_students():
    """Get all active students with their information for a specific user.
    
    Args:
    """
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        # Get only active student data for this owner
        c.execute("""
                             SELECT s.id, s.name, s.subject, s.email, s.phone, COALESCE(s.guardian, '') AS guardian, '' AS legacy_contact, s.active, s.book_loaned, s.device_loaned,
                 s.el, s.pi, s.v, s.day1, s.day1_time, s.day2, s.day2_time, s.subjects_json, s.subject_minutes_json, s.total_study_minutes,
                  s.photo_blob,
                  COALESCE(s.photo_mime, '') AS photo_mime,
                                    COALESCE(s.ind, 0) AS ind,
                                    COALESCE(s.schedule_json, '') AS schedule_json,
                                    COALESCE(s.day3, '') AS day3,
                                    COALESCE(s.day3_time, '') AS day3_time,
                                    COALESCE(s.day4, '') AS day4,
                                    COALESCE(s.day4_time, '') AS day4_time,
                                    COALESCE(s.day5, '') AS day5,
                                    COALESCE(s.day5_time, '') AS day5_time,
                                    COALESCE(s.day6, '') AS day6,
                                    COALESCE(s.day6_time, '') AS day6_time,
                                                                        COALESCE(s.student_identifier, '') AS student_identifier,
                                                                        COALESCE(s.secondary_email, '') AS secondary_email,
                                                                        COALESCE(s.secondary_phone, '') AS secondary_phone,
                                                                        COALESCE(s.secondary_guardian, '') AS secondary_guardian
            FROM students s
            WHERE s.active = 1
            ORDER BY s.name
        """, ())
        return c.fetchall()


def get_student(student_id):
    """Get a single student by ID, with ownership check.
    
    Args:
        student_id: Student ID to retrieve
    """
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        row = c.execute("""
             SELECT id,name,subject,email,phone,'' AS legacy_contact,active,book_loaned,device_loaned,
                 el,pi,v,day1,day2,day1_time,day2_time,subjects_json,subject_minutes_json,total_study_minutes,
                 photo_blob,
                 COALESCE(photo_mime,'') AS photo_mime,
                 COALESCE(schedule_json,'') AS schedule_json,
                 COALESCE(guardian,'') AS guardian,
                 COALESCE(ind,0) AS ind,
                 COALESCE(checkout_notify_enabled,1) AS checkout_notify_enabled,
                 COALESCE(day3,'') AS day3,
                 COALESCE(day3_time,'') AS day3_time,
                 COALESCE(day4,'') AS day4,
                 COALESCE(day4_time,'') AS day4_time,
                 COALESCE(day5,'') AS day5,
                 COALESCE(day5_time,'') AS day5_time,
                 COALESCE(day6,'') AS day6,
                 COALESCE(day6_time,'') AS day6_time,
                 COALESCE(student_identifier,'') AS student_identifier,
                 COALESCE(secondary_email,'') AS secondary_email,
                 COALESCE(secondary_phone,'') AS secondary_phone,
                 COALESCE(secondary_guardian,'') AS secondary_guardian
            FROM students WHERE id=?
        """, (student_id,)).fetchone()
        return row


def get_student_static_profile(student_id):
    """Get a single student by ID as a dictionary."""
    row = get_student(student_id)
    if not row:
        return None
    schedule_entries = _normalize_schedule_entries(
        row[21] if len(row) > 21 else '',
        (row[12] if len(row) > 12 else '', row[14] if len(row) > 14 else ''),
        (row[13] if len(row) > 13 else '', row[15] if len(row) > 15 else ''),
        (row[25] if len(row) > 25 else '', row[26] if len(row) > 26 else ''),
        (row[27] if len(row) > 27 else '', row[28] if len(row) > 28 else ''),
        (row[29] if len(row) > 29 else '', row[30] if len(row) > 30 else ''),
        (row[31] if len(row) > 31 else '', row[32] if len(row) > 32 else ''),
    )
    return {
        'id': row[0],
        'name': row[1],
        'subject': row[2],
        'email': row[3],
        'phone': row[4],
        'active': row[6],
        'book_loaned': row[7],
        'device_loaned': row[8],
        'el': row[9],
        'pi': row[10],
        'v': row[11],
        'day1': row[12],
        'day2': row[13],
        'day1_time': row[14],
        'day2_time': row[15],
        'day3': row[25] if len(row) > 25 else '',
        'day3_time': row[26] if len(row) > 26 else '',
        'day4': row[27] if len(row) > 27 else '',
        'day4_time': row[28] if len(row) > 28 else '',
        'day5': row[29] if len(row) > 29 else '',
        'day5_time': row[30] if len(row) > 30 else '',
        'day6': row[31] if len(row) > 31 else '',
        'day6_time': row[32] if len(row) > 32 else '',
        'subjects': json.loads(row[16] or '[]') if len(row) > 16 else ([row[2]] if row[2] else []),
        'subject_minutes': json.loads(row[17] or '[]') if len(row) > 17 else ([30] if row[2] else []),
        'total_study_minutes': int(row[18] or 30) if len(row) > 18 else 30,
        'photo_blob': _coerce_blob(row[19] if len(row) > 19 else None),
        'photo_mime': str(row[20] or '') if len(row) > 20 else '',
        'guardian': str(row[22] or '') if len(row) > 22 else '',
        'secondary_email': str(row[34] or '') if len(row) > 34 else '',
        'secondary_phone': str(row[35] or '') if len(row) > 35 else '',
        'secondary_guardian': str(row[36] or '') if len(row) > 36 else '',
        'photo_url': _photo_url(row[0], bool(_coerce_blob(row[19] if len(row) > 19 else None))),
        'checkout_notify_enabled': bool(row[24]) if len(row) > 24 else True,
        'student_identifier': str(row[33] or '') if len(row) > 33 else '',
        'schedule': schedule_entries,
    }


def get_student_photo(student_id):
    """Return the stored photo blob and metadata for a student."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            """
            SELECT photo_blob, COALESCE(photo_mime, '') AS photo_mime
            FROM students
            WHERE id=?
            """,
            (student_id,),
        ).fetchone()
        if not row:
            return None
        blob = _coerce_blob(row['photo_blob'])
        return {
            'photo_blob': blob,
            'photo_mime': str(row['photo_mime'] or '') or 'image/png',
        }


def set_student_photo(student_id, photo_blob=None, photo_mime='', photo_filename='', legacy_photo=''):
    """Set or clear a student's photo bytes and metadata.

    Legacy args ``photo_filename`` and ``legacy_photo`` are retained for
    backward compatibility with older call sites but are ignored.
    """
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "UPDATE students SET photo_blob=?, photo_mime=? WHERE id=?",
            (
                sqlite3.Binary(photo_blob) if photo_blob else None,
                photo_mime or '',
                student_id,
            ),
        )
        conn.commit()


def set_student_qr_code(student_id, qr_blob):
    """Store a student's QR code as BLOB in database."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "UPDATE students SET qr_code=? WHERE id=?",
            (sqlite3.Binary(qr_blob) if qr_blob else None, student_id),
        )
        conn.commit()

def set_checkout_notify_enabled(student_id, enabled):
    """Persist whether student should receive class checkout email notifications."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "UPDATE students SET checkout_notify_enabled=? WHERE id=?",
            (1 if bool(enabled) else 0, student_id),
        )
        conn.commit()

def get_student_qr_code(student_id):
    """Retrieve a student's QR code blob from database."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT qr_code FROM students WHERE id=?",
            (student_id,),
        ).fetchone()
        if row and row['qr_code']:
            return _coerce_blob(row['qr_code'])
    return None


def add_student(name, subject, email, phone, book_loaned=0, el=0, pi=0, v=0, ind=0, day1="", day2="", day1_time="", day2_time="", day3="", day3_time="", day4="", day4_time="", day5="", day5_time="", day6="", day6_time="", subjects=None, subject_minutes=None, schedule_json="", guardian="", student_identifier="", secondary_email="", secondary_phone="", secondary_guardian=""):
    """Add a new student to the database and automatically generate QR code.
    
    Args:
    """
    subjects_list, minutes_list, total_minutes = normalize_subject_entries(
        subjects if subjects is not None else [subject],
        subject_minutes if subject_minutes is not None else [30],
    )
    primary_subject = subjects_list[0]

    with sqlite3.connect(DB_PATH) as conn:
        _advance_student_id_sequence_past_repeating_digits(conn)
        c = conn.cursor()
        c.execute("""INSERT INTO students
            (name,student_identifier,subject,subjects_json,subject_minutes_json,total_study_minutes,email,phone,guardian,secondary_email,secondary_phone,secondary_guardian,active,book_loaned,el,pi,v,ind,day1,day2,day1_time,day2_time,day3,day3_time,day4,day4_time,day5,day5_time,day6,day6_time,schedule_json)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                name,
                normalize_student_identifier(student_identifier),
                primary_subject,
                json.dumps(subjects_list),
                json.dumps(minutes_list),
                total_minutes,
                email,
                phone,
                str(guardian or '').strip(),
                str(secondary_email or '').strip(),
                str(secondary_phone or '').strip(),
                str(secondary_guardian or '').strip(),
                1,
                int(bool(book_loaned)),
                int(bool(el)),
                int(bool(pi)),
                int(bool(v)),
                int(bool(ind)),
                day1,
                day2,
                day1_time,
                day2_time,
                day3,
                day3_time,
                day4,
                day4_time,
                day5,
                day5_time,
                day6,
                day6_time,
                schedule_json,
            ))
        student_id = c.lastrowid
        conn.commit()
    
    # Automatically generate QR code for the new student and store in DB only if not exists
    try:
        existing_qr = get_student_qr_code(student_id)
        if not existing_qr:
            unique_token = issue_unique_qr_token("STU", "student", student_id)
            qr_data = f"ID:{student_id}\nName:{name}\nUID:{unique_token}"
            qr_blob = qr_generator.generate_qr_bytes(qr_data)
            set_student_qr_code(student_id, qr_blob)
    except Exception as e:
        print(f"Warning: Failed to generate QR code for student {student_id}: {e}")
    return student_id


def _is_missing_text(value):
    return not str(value or '').strip()


def _find_existing_student_id_for_vcf(
    conn,
    student_identifier='',
    email='',
    name='',
    phone='',
    match_on_identifier=True,
    match_on_email=True,
):
    """Find an existing student for VCF upsert using priority: ID > email > name+phone."""
    normalized_identifier = normalize_student_identifier(student_identifier)
    email = str(email or '').strip()
    name = str(name or '').strip()
    phone = str(phone or '').strip()

    if match_on_identifier and normalized_identifier:
        row = conn.execute(
            "SELECT id FROM students WHERE LOWER(COALESCE(student_identifier,'')) = LOWER(?) LIMIT 1",
            (normalized_identifier,),
        ).fetchone()
        if row:
            return int(row[0])

    if match_on_email and email:
        row = conn.execute(
            "SELECT id FROM students WHERE LOWER(COALESCE(email,'')) = LOWER(?) LIMIT 1",
            (email,),
        ).fetchone()
        if row:
            return int(row[0])

    if name and phone:
        row = conn.execute(
            "SELECT id FROM students WHERE LOWER(COALESCE(name,'')) = LOWER(?) AND LOWER(COALESCE(phone,'')) = LOWER(?) LIMIT 1",
            (name, phone),
        ).fetchone()
        if row:
            return int(row[0])

    return None


def upsert_student_from_vcf_contact(
    student_name,
    email='',
    phone='',
    guardian='',
    student_identifier='',
    subjects=None,
    subject_minutes=None,
    match_on_email=True,
    match_on_identifier=True,
):
    """Safely upsert one student from VCF data.

    Matching priority: student_identifier > email > name+phone.
    Updates only missing target fields (email/phone/guardian/student_identifier).
    """
    name = str(student_name or '').strip()
    if not name:
        return {'action': 'skipped', 'reason': 'missing_name'}

    email = str(email or '').strip()
    phone = str(phone or '').strip()
    guardian = str(guardian or '').strip()
    normalized_identifier = normalize_student_identifier(student_identifier)

    with sqlite3.connect(DB_PATH) as conn:
        existing_id = _find_existing_student_id_for_vcf(
            conn,
            student_identifier=normalized_identifier,
            email=email,
            name=name,
            phone=phone,
            match_on_identifier=match_on_identifier,
            match_on_email=match_on_email,
        )

        if existing_id:
            row = conn.execute(
                """
                SELECT
                    COALESCE(email, ''),
                    COALESCE(phone, ''),
                    COALESCE(guardian, ''),
                    COALESCE(student_identifier, ''),
                    COALESCE(subject, ''),
                    COALESCE(subjects_json, '[]'),
                    COALESCE(subject_minutes_json, '[]')
                FROM students
                WHERE id = ?
                LIMIT 1
                """,
                (existing_id,),
            ).fetchone()

            if not row:
                return {'action': 'skipped', 'reason': 'existing_not_found'}

            existing_email, existing_phone, existing_guardian, existing_identifier, existing_subject, existing_subjects_json, existing_subject_minutes_json = row
            updates = {}

            if email and _is_missing_text(existing_email):
                updates['email'] = email
            if phone and _is_missing_text(existing_phone):
                updates['phone'] = phone
            if guardian and _is_missing_text(existing_guardian):
                updates['guardian'] = guardian
            if normalized_identifier and _is_missing_text(existing_identifier):
                updates['student_identifier'] = normalized_identifier

            subjects_list, minutes_list, total_minutes = normalize_subject_entries(
                subjects if subjects is not None else [existing_subject] if existing_subject else [],
                subject_minutes if subject_minutes is not None else [30] if subjects else [],
            )
            if subjects and _is_missing_text(existing_subject) and str(existing_subjects_json or '').strip() in ('', '[]'):
                updates['subject'] = subjects_list[0]
                updates['subjects_json'] = json.dumps(subjects_list)
                updates['subject_minutes_json'] = json.dumps(minutes_list)
                updates['total_study_minutes'] = total_minutes

            if not updates:
                return {'action': 'skipped', 'student_id': existing_id, 'reason': 'already_complete'}

            set_clause = ', '.join([f"{key}=?" for key in updates.keys()])
            params = list(updates.values()) + [existing_id]
            conn.execute(f"UPDATE students SET {set_clause} WHERE id=?", params)
            conn.commit()
            return {
                'action': 'updated',
                'student_id': existing_id,
                'updated_fields': list(updates.keys()),
            }

    student_id = add_student(
        name,
        (subjects[0] if subjects else "Math"),
        email,
        phone,
        book_loaned=0,
        el=0,
        pi=1,
        v=0,
        ind=0,
        guardian=guardian,
        student_identifier=normalized_identifier,
        subjects=subjects,
        subject_minutes=subject_minutes,
    )
    return {'action': 'added', 'student_id': student_id}



def update_student(sid, name, email, phone, subject="", book_loaned=0, el=0, pi=0, v=0, ind=0, day1="", day2="", day1_time="", day2_time="", day3="", day3_time="", day4="", day4_time="", day5="", day5_time="", day6="", day6_time="", subjects=None, subject_minutes=None, schedule_json="", guardian="", student_identifier="", secondary_email="", secondary_phone="", secondary_guardian=""):
    """Update an existing student's information with ownership check."""
    subjects_list, minutes_list, total_minutes = normalize_subject_entries(
        subjects if subjects is not None else [subject],
        subject_minutes if subject_minutes is not None else [30],
    )
    primary_subject = subjects_list[0]

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""UPDATE students SET name=?,student_identifier=?,subject=?,subjects_json=?,subject_minutes_json=?,total_study_minutes=?,email=?,phone=?,guardian=?,secondary_email=?,secondary_phone=?,secondary_guardian=?,book_loaned=?,el=?,pi=?,v=?,ind=?,day1=?,day2=?,day1_time=?,day2_time=?,day3=?,day3_time=?,day4=?,day4_time=?,day5=?,day5_time=?,day6=?,day6_time=?,schedule_json=? WHERE id=?""",
                  (
                      name,
                      normalize_student_identifier(student_identifier),
                      primary_subject,
                      json.dumps(subjects_list),
                      json.dumps(minutes_list),
                      total_minutes,
                      email,
                      phone,
                      str(guardian or '').strip(),
                      str(secondary_email or '').strip(),
                      str(secondary_phone or '').strip(),
                      str(secondary_guardian or '').strip(),
                      int(bool(book_loaned)),
                      int(bool(el)),
                      int(bool(pi)),
                      int(bool(v)),
                      int(bool(ind)),
                      day1,
                      day2,
                      day1_time,
                      day2_time,
                      day3,
                      day3_time,
                      day4,
                      day4_time,
                      day5,
                      day5_time,
                      day6,
                      day6_time,
                      schedule_json,
                      sid,
                  ))
        conn.commit()


def delete_student(sid):
    """Soft delete: mark student as inactive instead of hard delete with ownership check."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("UPDATE students SET active=0 WHERE id=?", (sid,))
        conn.commit()


def permanent_delete_student(sid):
    """Permanently delete student from database (hard delete) with ownership check."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("DELETE FROM students WHERE id=?", (sid,))
        conn.commit()


def get_deleted_students():
    """Get all deleted/inactive students for a specific user."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT s.id, s.name, s.subject, s.email, s.phone, '' AS legacy_contact, s.active, s.book_loaned, s.device_loaned,
                   s.el, s.pi, s.v, s.day1, s.day1_time, s.day2, s.day2_time
            FROM students s
            WHERE s.active = 0
            ORDER BY s.name
        """, ())
        return c.fetchall()


def reactivate_student(sid):
    """Reactivate a deleted/inactive student with ownership check."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("UPDATE students SET active=1 WHERE id=?", (sid,))
        conn.commit()


def _xlsx_key(value):
    """Normalize Excel headers and matching names for safe comparisons."""
    return ''.join(ch for ch in str(value or '').strip().casefold() if ch.isalnum())


def _xlsx_text(value):
    """Convert an Excel cell to trimmed text without turning blanks into 'None'."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _student_name_match_key(first_name, last_name):
    """Match on first name plus complete last-name initial."""
    first = ' '.join(_xlsx_text(first_name).casefold().split())
    last = ' '.join(_xlsx_text(last_name).casefold().split())
    return f"{first}|{last[:1]}" if first and last else ''


def _xlsx_header(row, *aliases):
    values = {_xlsx_key(alias) for alias in aliases}
    for key, value in row.items():
        if _xlsx_key(key) in values:
            return _xlsx_text(value)
    return ''


def import_students_from_xlsx(file_path):
    """Insert new students from Excel while protecting existing matches."""
    if not os.path.exists(file_path):
        return {"error": "The selected Excel file could not be found."}

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return {"error": "Excel import requires the openpyxl package.", "detail": str(exc)}

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if not raw_headers:
            return {"error": "The Excel worksheet is empty."}

        headers = [_xlsx_text(value) for value in raw_headers]
        header_keys = {_xlsx_key(header) for header in headers if header}
        first_aliases = {'studentfirstname', 'firstname', 'studentfirst'}
        last_aliases = {'studentlastname', 'lastname', 'studentlast', 'surname'}
        if not header_keys.intersection(first_aliases) or not header_keys.intersection(last_aliases):
            return {"error": "The Excel file must contain Student First Name and Student Last Name columns."}

        incoming = []
        invalid = []
        for row_number, values in enumerate(rows, start=2):
            row = dict(zip(headers, values))
            first_name = _xlsx_header(row, 'Student First Name', 'First Name', 'Student First')
            last_name = _xlsx_header(row, 'Student Last Name', 'Last Name', 'Student Last', 'Surname')
            if not first_name and not last_name and not any(values):
                continue
            if not first_name or not last_name:
                invalid.append({"row": row_number, "reason": "Student first and last name are required."})
                continue

            mother_first = _xlsx_header(row, 'Mother First Name', 'Mother First')
            father_first = _xlsx_header(row, 'Father First Name', 'Father First')
            mother = {
                'name': f"Ms. {mother_first}" if mother_first else '',
                'phone': _xlsx_header(row, 'Mother cell phone', 'Mother Cell Phone', 'Mother Phone'),
                'email': _xlsx_header(row, 'Mother Email'),
            }
            father = {
                'name': f"Mr. {father_first}" if father_first else '',
                'phone': _xlsx_header(row, 'Father cell phone', 'Father Cell Phone', 'Father Phone'),
                'email': _xlsx_header(row, 'Father Email'),
            }
            incoming.append({
                'row': row_number,
                'first_name': first_name,
                'last_name': last_name,
                'name': f"{first_name} {last_name}".strip(),
                'match_key': _student_name_match_key(first_name, last_name),
                'student_identifier': normalize_student_identifier(
                    _xlsx_header(row, 'Student ID', 'Student Identifier', 'StudentID')
                ),
                'mother': mother,
                'father': father,
            })
        workbook.close()
    except Exception as exc:
        return {"error": f"Could not read the Excel workbook: {exc}"}

    result = {'added': 0, 'skipped_existing': 0, 'skipped_ambiguous': 0,
              'skipped_duplicate_input': 0, 'invalid': len(invalid), 'rows': invalid}
    incoming_keys = {}
    new_students_for_qr = []
    for item in incoming:
        incoming_keys.setdefault(item['match_key'], []).append(item)

    with sqlite3.connect(DB_PATH) as conn:
        existing = {}
        for student_id, name, stored_last_name in conn.execute(
            "SELECT id, COALESCE(name, ''), COALESCE(last_name, '') FROM students"
        ):
            parts = str(name or '').strip().split()
            derived_last = stored_last_name or (parts[-1] if len(parts) >= 2 else '')
            key = _student_name_match_key(parts[0] if parts else '', derived_last)
            if key:
                existing.setdefault(key, []).append((student_id, name))

        for key, items in incoming_keys.items():
            if len(items) > 1:
                for item in items:
                    result['skipped_duplicate_input'] += 1
                    result['rows'].append({'row': item['row'], 'student': item['name'],
                                           'action': 'skipped_duplicate_input',
                                           'reason': 'Duplicate first-name and last-initial key in workbook.'})
                continue

            item = items[0]
            matches = existing.get(key, [])
            if len(matches) == 1:
                result['skipped_existing'] += 1
                result['rows'].append({'row': item['row'], 'student': item['name'],
                                       'action': 'skipped_existing',
                                       'reason': f"Protected existing match: {matches[0][1]} (ID {matches[0][0]})."})
                continue
            if len(matches) > 1:
                result['skipped_ambiguous'] += 1
                result['rows'].append({'row': item['row'], 'student': item['name'],
                                       'action': 'skipped_ambiguous',
                                       'reason': 'Multiple existing students share the same first name and last initial.'})
                continue

            mother = item['mother']
            father = item['father']
            primary = mother if mother['name'] else father
            secondary = father if mother['name'] and father['name'] else {'name': '', 'phone': '', 'email': ''}
            _advance_student_id_sequence_past_repeating_digits(conn)
            cursor = conn.execute(
                """
                INSERT INTO students (
                    name, last_name, student_identifier, subject, subjects_json,
                    subject_minutes_json, total_study_minutes, email, phone, guardian,
                    secondary_email, secondary_phone, secondary_guardian, active, pi
                ) VALUES (?, ?, ?, 'Math', '[\"Math\"]', '[30]', 30, ?, ?, ?, ?, ?, ?, 1, 1)
                """,
                (item['name'], item['last_name'], item['student_identifier'],
                 primary['email'], primary['phone'], primary['name'],
                 secondary['email'], secondary['phone'], secondary['name']),
            )
            student_id = cursor.lastrowid
            new_students_for_qr.append((student_id, item['name']))
            result['added'] += 1
            result['rows'].append({'row': item['row'], 'student': item['name'], 'action': 'added', 'student_id': student_id})
            existing.setdefault(key, []).append((student_id, item['name']))
        conn.commit()

    # Generate QR codes only after the import transaction closes.  The QR
    # registry uses its own SQLite connection and must not nest inside this one.
    for student_id, student_name in new_students_for_qr:
        try:
            unique_token = issue_unique_qr_token("STU", "student", student_id)
            qr_data = f"ID:{student_id}\nName:{student_name}\nUID:{unique_token}"
            qr_blob = qr_generator.generate_qr_bytes(qr_data)
            with sqlite3.connect(DB_PATH) as qr_conn:
                qr_conn.execute("UPDATE students SET qr_code=? WHERE id=?", (sqlite3.Binary(qr_blob), student_id))
        except Exception as qr_err:
            print(f"Warning: Failed to generate QR for new student {student_id}: {qr_err}")
    return result

def export_csv(path):
    """Export active students in the same shape used by the student edit form CSV import."""
    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute(
            """
            SELECT
                name,
                COALESCE(email, ''),
                COALESCE(phone, ''),
                COALESCE(guardian, ''),
                COALESCE(student_identifier, ''),
                COALESCE(subjects_json, '[]'),
                COALESCE(subject, ''),
                el,
                pi,
                v,
                COALESCE(ind, 0)
            FROM students
            WHERE active = 1
            ORDER BY name
            """
        ).fetchall()

    headers = [
        "name",
        "email",
        "phone",
        "guardian",
        "student_id",
        "M",
        "R",
        "W",
        "classification",
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            name, email, phone, guardian, student_identifier = row[0], row[1], row[2], row[3], row[4]
            subjects_json, legacy_subject = row[5], row[6]
            el, pi, v, ind = int(bool(row[7])), int(bool(row[8])), int(bool(row[9])), int(bool(row[10]))

            try:
                subjects = [str(s).strip() for s in json.loads(subjects_json or '[]') if str(s or '').strip()]
            except (TypeError, ValueError):
                subjects = []
            if not subjects and legacy_subject:
                subjects = [str(legacy_subject).strip()]

            normalized_subjects = {_normalize_subject_name(s).lower() for s in subjects}
            m_flag = 'x' if 'math' in normalized_subjects else ''
            r_flag = 'x' if 'reading' in normalized_subjects else ''
            w_flag = 'x' if 'writing' in normalized_subjects else ''

            classification = _classification_label(el=el, pi=pi, v=v, ind=ind)

            writer.writerow([
                name,
                email,
                phone,
                guardian,
                student_identifier,
                m_flag,
                r_flag,
                w_flag,
                classification,
            ])


def get_students_mailing_list_rows():
    """Return active student rows for mailing export (name, email)."""
    with sqlite3.connect(DB_PATH) as conn:
        return conn.execute(
            """
            SELECT COALESCE(name, ''), COALESCE(email, '')
            FROM students
            WHERE active = 1
              AND TRIM(COALESCE(email, '')) <> ''
            ORDER BY name
            """
        ).fetchall()


def get_students_badge_payload():
    """Return active student payload used by badge PDF rendering."""
    with sqlite3.connect(DB_PATH) as conn:
        return conn.execute(
            """
            SELECT
                id,
                COALESCE(name, ''),
                COALESCE(student_identifier, ''),
                photo_blob,
                COALESCE(photo_mime, ''),
                qr_code
            FROM students
            WHERE active = 1
            ORDER BY name
            """
        ).fetchall()


def find_duplicates_by_name(name):
    """Find all students with a given name (case-insensitive, whitespace-trimmed)."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT id, name, email, phone, subject, day1, day1_time, day2, day2_time
            FROM students
            WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
            ORDER BY id
        """, (name,))
        return c.fetchall()


def get_duplicate_names():
    """Get all student names that appear more than once in the student list."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT LOWER(TRIM(name)) as name_key, COUNT(*) as count
            FROM students
            WHERE active = 1
            GROUP BY LOWER(TRIM(name))
            HAVING COUNT(*) > 1
            ORDER BY count DESC, name_key
        """, ())
        return c.fetchall()


def get_duplicate_name_count():
    """Return how many distinct duplicate names exist among active students."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        row = c.execute(
            """
            SELECT COUNT(*)
            FROM (
                SELECT LOWER(TRIM(name))
                FROM students
                WHERE active = 1
                GROUP BY LOWER(TRIM(name))
                HAVING COUNT(*) > 1
            )
            """,
            (),
        ).fetchone()
        return int(row[0] or 0) if row else 0


def get_duplicate_summary():
    """Get a detailed summary of all duplicate names with their student information."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        rows = c.execute(
            """
            WITH duplicate_names AS (
                SELECT LOWER(TRIM(name)) AS name_key, COUNT(*) AS dup_count
                FROM students
                WHERE active = 1
                GROUP BY LOWER(TRIM(name))
                HAVING COUNT(*) > 1
            )
            SELECT
                d.name_key,
                d.dup_count,
                s.id,
                s.name,
                s.email,
                s.phone,
                s.subject,
                s.day1,
                s.day1_time,
                s.day2,
                s.day2_time,
                s.el,
                s.pi,
                s.v
            FROM duplicate_names d
            JOIN students s
              ON LOWER(TRIM(s.name)) = d.name_key
             AND s.active = 1
            ORDER BY d.dup_count DESC, d.name_key, s.id
            """,
            (),
        ).fetchall()

    summary = []
    current_key = None
    current_block = None

    for row in rows:
        name_key = row[0]
        dup_count = row[1]
        if current_key != name_key:
            current_key = name_key
            current_block = {
                'name': row[3],
                'count': dup_count,
                'students': [],
            }
            summary.append(current_block)

        current_block['students'].append(
            {
                'id': row[2],
                'name': row[3],
                'email': row[4],
                'phone': row[5],
                'subject': row[6],
                'day1': row[7],
                'day1_time': row[8],
                'day2': row[9],
                'day2_time': row[10],
                'el': row[11],
                'pi': row[12],
                'v': row[13],
            }
        )

    return summary


def has_duplicate_names():
    """Check if there are any duplicate names in the active student list."""
    duplicates = get_duplicate_names()
    return len(duplicates) > 0